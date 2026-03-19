import pandas as pd
import fitz  # PyMuPDF
import os
import re
import unicodedata
from typing import Dict, Any, Union
import json
import io

def _score_text_quality(text: str) -> dict:
    """텍스트 품질을 0~100점으로 평가합니다."""
    if not text or not text.strip():
        return {"score": 0, "details": {}}

    stripped = text.strip()
    total_chars = len(stripped)

    # 1. 텍스트 밀도 (20점): 공백/줄바꿈 제외 실제 내용 비율
    non_space = len(re.sub(r'\s', '', stripped))
    density_ratio = non_space / total_chars if total_chars > 0 else 0
    density_score = min(20, int(density_ratio * 25))

    # 2. 언어 일관성 (20점)
    korean_chars = len(re.findall(r'[\uAC00-\uD7A3]', stripped))
    english_chars = len(re.findall(r'[a-zA-Z]', stripped))
    alphanum = korean_chars + english_chars
    if alphanum > 0:
        ko_ratio = korean_chars / alphanum
        en_ratio = english_chars / alphanum
        # 한 언어가 30% 이상 차지하면 일관성 있음
        if ko_ratio > 0.3 or en_ratio > 0.5:
            lang_score = 20
        else:
            lang_score = 10
    else:
        lang_score = 0

    # 3. 구조 보존도 (20점): 줄바꿈 패턴, 단락 구분 존재
    lines = stripped.split('\n')
    non_empty_lines = [l for l in lines if l.strip()]
    if len(lines) > 1 and len(non_empty_lines) >= 3:
        struct_score = 20
    elif len(non_empty_lines) >= 1:
        struct_score = 10
    else:
        struct_score = 0

    # 4. 인코딩 건강도 (20점): 제어문자·알 수 없는 유니코드 비율 (줄바꿈/탭은 제외)
    control_chars = sum(1 for c in stripped if unicodedata.category(c) in ('Cc', 'Cs', 'Co') and c not in '\n\r\t')
    control_ratio = control_chars / total_chars if total_chars > 0 else 0
    if control_ratio < 0.01:
        enc_score = 20
    elif control_ratio < 0.05:
        enc_score = 10
    else:
        enc_score = 0

    # 5. 내용 의미밀도 (20점): 단어 평균 길이 2~15, 반복 패턴 없음
    words = re.findall(r'\S+', stripped)
    if words:
        avg_word_len = sum(len(w) for w in words) / len(words)
        # 단어 다양성 체크 (상위 5개 단어가 전체의 50% 미만이면 좋음)
        from collections import Counter
        word_counts = Counter(words)
        top5_count = sum(c for _, c in word_counts.most_common(5))
        diversity = 1 - (top5_count / len(words)) if len(words) > 5 else 0.5
        if 2 <= avg_word_len <= 15 and diversity > 0.5:
            semantic_score = 20
        elif 1.5 <= avg_word_len <= 20:
            semantic_score = 10
        else:
            semantic_score = 5
    else:
        semantic_score = 0

    total = density_score + lang_score + struct_score + enc_score + semantic_score
    return {
        "score": total,
        "details": {
            "density": density_score,
            "language": lang_score,
            "structure": struct_score,
            "encoding": enc_score,
            "semantic": semantic_score
        }
    }


class UniversalFileAnalyzer:
    """다양한 파일 형식을 지원하는 범용 파일 분석기"""
    
    def __init__(self, use_gemini_pdf: bool = False):
        """
        Args:
            use_gemini_pdf: PDF 처리 시 Gemini 네이티브 방식을 사용할지 여부
                           False면 PyMuPDF로 빠른 텍스트 추출만 수행
        """
        self.use_gemini_pdf = use_gemini_pdf
        self.supported_formats = {
            "pdf": self._analyze_pdf,
            "xlsx": self._analyze_excel,
            "xls": self._analyze_excel,
            "csv": self._analyze_csv,
            "txt": self._analyze_text,
            "json": self._analyze_json,
            "docx": self._analyze_docx
        }
    
    def analyze_file(self, file_path: str, file_type: str) -> Dict[str, Any]:
        """파일을 분석하고 텍스트를 추출합니다."""
        if file_type.lower() not in self.supported_formats:
            return {
                "success": False,
                "error": f"지원하지 않는 파일 형식입니다: {file_type}",
                "supported_formats": list(self.supported_formats.keys())
            }
        
        try:
            return self.supported_formats[file_type.lower()](file_path)
        except Exception as e:
            return {
                "success": False,
                "error": f"파일 분석 중 오류 발생: {str(e)}"
            }
    
    # 텍스트 최대 길이 (약 25만 토큰 수준, 분석 품질·비용 균형)
    MAX_TEXT_CHARS = 200_000

    def analyze_file_from_bytes(self, file_bytes: bytes, file_type: str, file_name: str = "") -> Dict[str, Any]:
        """바이트 데이터로부터 파일을 분석하고 텍스트를 추출합니다."""
        if file_type.lower() not in self.supported_formats:
            return {
                "success": False,
                "error": f"지원하지 않는 파일 형식입니다: {file_type}",
                "supported_formats": list(self.supported_formats.keys())
            }

        try:
            result = self._analyze_from_bytes(file_bytes, file_type.lower(), file_name)
        except Exception as e:
            return {
                "success": False,
                "error": f"파일 분석 중 오류 발생: {str(e)}"
            }

        # 대용량 텍스트 자르기
        if result.get("success") and "text" in result:
            text = result["text"]
            if len(text) > self.MAX_TEXT_CHARS:
                result["text"] = text[:self.MAX_TEXT_CHARS]
                result["truncated"] = True
                result["original_char_count"] = len(text)
                result["char_count"] = self.MAX_TEXT_CHARS
                print(f"[FileAnalyzer] 텍스트 길이 제한 적용: {len(text):,} → {self.MAX_TEXT_CHARS:,}자")
            else:
                result["truncated"] = False

        return result
    
    def _analyze_from_bytes(self, file_bytes: bytes, file_type: str, file_name: str) -> Dict[str, Any]:
        """바이트 데이터로부터 파일 분석"""
        if file_type == "pdf":
            return self._analyze_pdf_from_bytes(file_bytes)
        elif file_type in ["xlsx", "xls"]:
            return self._analyze_excel_from_bytes(file_bytes)
        elif file_type == "csv":
            return self._analyze_csv_from_bytes(file_bytes)
        elif file_type == "txt":
            return self._analyze_text_from_bytes(file_bytes)
        elif file_type == "json":
            return self._analyze_json_from_bytes(file_bytes)
        elif file_type == "docx":
            return self._analyze_docx_from_bytes(file_bytes)
        else:
            return {
                "success": False,
                "error": f"지원하지 않는 파일 형식입니다: {file_type}"
            }
    
    def _analyze_pdf_from_bytes(self, file_bytes: bytes) -> Dict[str, Any]:
        """PDF 바이트 데이터 분석 (품질 점수 기반 파이프라인)"""
        candidate_text = None
        candidate_score = 0
        candidate_method = None
        page_count = 0
        is_scanned = False

        # Step 1: pymupdf4llm으로 Markdown 추출 시도
        try:
            import pymupdf4llm
            md_text = pymupdf4llm.to_markdown(fitz.open(stream=file_bytes, filetype="pdf"))
            quality = _score_text_quality(md_text)
            score = quality["score"]
            print(f"[PDF] pymupdf4llm 품질 점수: {score}")
            if score >= 60:
                # 60+ 즉시 채택
                doc = fitz.open(stream=file_bytes, filetype="pdf")
                page_count = len(doc)
                doc.close()
                text = md_text.strip()
                return {
                    "success": True,
                    "file_type": "pdf",
                    "text": text,
                    "page_count": page_count,
                    "word_count": len(text.split()),
                    "char_count": len(text),
                    "preview": text[:500] + "..." if len(text) > 500 else text,
                    "method": "pymupdf4llm",
                    "quality_score": score,
                    "is_scanned": False
                }
            elif score >= 40:
                candidate_text = md_text
                candidate_score = score
                candidate_method = "pymupdf4llm"
        except Exception as e:
            print(f"[PDF] pymupdf4llm 실패: {e}")

        # Step 2: PyMuPDF page.get_text() 방식
        try:
            doc = fitz.open(stream=file_bytes, filetype="pdf")
            page_count = len(doc)
            text_parts = []
            page_text_lengths = []
            has_images = False

            for page_num in range(page_count):
                try:
                    page = doc[page_num]
                    page_text = page.get_text()
                    page_text_lengths.append(len(page_text.strip()))
                    if page_text:
                        text_parts.append(page_text)
                    # 이미지 존재 여부 확인
                    if page.get_images():
                        has_images = True
                except Exception as page_error:
                    print(f"페이지 {page_num + 1} 처리 중 오류: {page_error}")
                    page_text_lengths.append(0)

            doc.close()
            pymupdf_text = "\n".join(text_parts).strip()

            # 스캔 PDF 감지: 페이지당 평균 텍스트 < 50자 AND 이미지 존재
            avg_text_per_page = (sum(page_text_lengths) / len(page_text_lengths)) if page_text_lengths else 0
            if avg_text_per_page < 50 and has_images:
                is_scanned = True

            if pymupdf_text:
                quality = _score_text_quality(pymupdf_text)
                score = quality["score"]
                print(f"[PDF] PyMuPDF 품질 점수: {score}")
                if candidate_text is None or score > candidate_score:
                    candidate_text = pymupdf_text
                    candidate_score = score
                    candidate_method = "pymupdf"

        except Exception as e:
            print(f"[PDF] PyMuPDF 실패: {e}")

        # Step 3: 스캔 PDF + Gemini API 키 있으면 Gemini 폴백
        if is_scanned and (self.use_gemini_pdf or os.environ.get("GOOGLE_API_KEY")):
            try:
                from pdf_analyzer import extract_text_with_gemini_pdf
                result = extract_text_with_gemini_pdf(
                    pdf_path=file_bytes,
                    prompt="이 PDF 문서의 모든 텍스트 내용을 추출하고 구조화해주세요."
                )
                if result.get("success"):
                    text = result.get("text", "")
                    quality = _score_text_quality(text)
                    if quality["score"] > candidate_score:
                        candidate_text = text
                        candidate_score = quality["score"]
                        candidate_method = "gemini"
            except Exception as e:
                print(f"[PDF] Gemini 폴백 실패: {e}")

        # 최종 결과 반환
        if candidate_text and candidate_score >= 1:
            text = candidate_text.strip()
            return {
                "success": True,
                "file_type": "pdf",
                "text": text,
                "page_count": page_count,
                "word_count": len(text.split()),
                "char_count": len(text),
                "preview": text[:500] + "..." if len(text) > 500 else text,
                "method": candidate_method,
                "quality_score": candidate_score,
                "is_scanned": is_scanned
            }

        # 아무것도 추출 못한 경우
        return {
            "success": False,
            "error": "PDF에서 텍스트를 추출할 수 없습니다. 이미지 기반 PDF이거나 텍스트가 없는 PDF일 수 있습니다.",
            "is_scanned": is_scanned,
            "page_count": page_count
        }
    
    def _excel_fill_merged_cells(self, ws) -> None:
        """openpyxl 워크시트의 병합 셀을 상단 좌측 값으로 채웁니다."""
        import copy
        merge_ranges = list(ws.merged_cells.ranges)
        for merge_range in merge_ranges:
            min_row, min_col = merge_range.min_row, merge_range.min_col
            top_left_value = ws.cell(min_row, min_col).value
            ws.unmerge_cells(str(merge_range))
            for row in ws.iter_rows(
                min_row=merge_range.min_row, max_row=merge_range.max_row,
                min_col=merge_range.min_col, max_col=merge_range.max_col
            ):
                for cell in row:
                    cell.value = top_left_value

    def _analyze_excel_from_bytes(self, file_bytes: bytes) -> Dict[str, Any]:
        """Excel 바이트 데이터 분석 (Markdown 테이블 변환 + 병합 셀 처리)"""
        try:
            import openpyxl

            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            sheets_data = {}
            all_text = ""

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # 병합 셀 → 상단 좌측 값으로 채우기
                self._excel_fill_merged_cells(ws)

                # 워크시트 → DataFrame
                data = list(ws.values)
                if not data:
                    continue

                # 첫 행을 컬럼명으로 사용, 나머지를 데이터로
                headers = [str(h) if h is not None else "" for h in data[0]]
                rows = data[1:]
                df = pd.DataFrame(rows, columns=headers)

                # 완전히 빈 행/열 제거
                df.dropna(how="all", inplace=True)
                df.dropna(axis=1, how="all", inplace=True)
                df.reset_index(drop=True, inplace=True)

                if df.empty:
                    continue

                # Markdown pipe table 변환 (tabulate는 pymupdf4llm 의존성으로 설치됨)
                try:
                    sheet_text = df.to_markdown(index=False)
                except Exception:
                    sheet_text = df.to_string(index=False)

                sheets_data[sheet_name] = {
                    "shape": df.shape,
                    "columns": df.columns.tolist(),
                    "text": sheet_text
                }
                all_text += f"\n## {sheet_name}\n\n{sheet_text}\n"

            if not all_text.strip():
                return {
                    "success": False,
                    "error": "Excel 파일에서 내용을 추출할 수 없습니다. 빈 파일이거나 지원하지 않는 형식일 수 있습니다."
                }

            all_text = all_text.strip()
            return {
                "success": True,
                "file_type": "excel",
                "text": all_text,
                "sheets": sheets_data,
                "sheet_count": len(sheets_data),
                "sheet_names": list(sheets_data.keys()),
                "word_count": len(all_text.split()),
                "char_count": len(all_text),
                "preview": all_text[:500] + "..." if len(all_text) > 500 else all_text
            }

        except Exception as e:
            return {
                "success": False,
                "error": f"Excel 분석 오류: {str(e)}"
            }
    
    def _analyze_csv_from_bytes(self, file_bytes: bytes) -> Dict[str, Any]:
        """CSV 바이트 데이터 분석 (인코딩 자동 감지 + 구분자 자동 감지)"""
        try:
            # 인코딩 순차 시도 (한국 행정 데이터는 CP949가 많음)
            encodings = ["utf-8", "cp949", "euc-kr", "utf-8-sig", "latin-1"]
            df = None
            used_encoding = None

            for enc in encodings:
                try:
                    # sep=None + engine='python' → 구분자 자동 감지 (쉼표/세미콜론/탭 등)
                    df = pd.read_csv(io.BytesIO(file_bytes), encoding=enc, sep=None, engine="python")
                    used_encoding = enc
                    break
                except (UnicodeDecodeError, Exception):
                    continue

            if df is None:
                return {
                    "success": False,
                    "error": "CSV 파일을 읽을 수 없습니다. 인코딩 또는 형식을 확인해주세요."
                }

            # 데이터프레임을 텍스트로 변환
            csv_text = df.to_string(index=False)

            return {
                "success": True,
                "file_type": "csv",
                "text": csv_text,
                "shape": df.shape,
                "columns": df.columns.tolist(),
                "data_types": df.dtypes.to_dict(),
                "encoding": used_encoding,
                "word_count": len(csv_text.split()),
                "char_count": len(csv_text),
                "preview": csv_text[:500] + "..." if len(csv_text) > 500 else csv_text
            }

        except Exception as e:
            return {
                "success": False,
                "error": f"CSV 분석 오류: {str(e)}"
            }
    
    def _analyze_text_from_bytes(self, file_bytes: bytes) -> Dict[str, Any]:
        """텍스트 바이트 데이터 분석"""
        try:
            # 다양한 인코딩으로 시도
            encodings = ["utf-8", "cp949", "euc-kr", "latin-1"]
            text = ""
            
            for encoding in encodings:
                try:
                    text = file_bytes.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            
            if not text:
                raise ValueError("텍스트 파일을 읽을 수 없습니다.")
            
            return {
                "success": True,
                "file_type": "text",
                "text": text.strip(),
                "word_count": len(text.split()),
                "char_count": len(text),
                "line_count": len(text.split("\n")),
                "preview": text[:500] + "..." if len(text) > 500 else text
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": f"텍스트 파일 분석 오류: {str(e)}"
            }
    
    def _json_schema_summary(self, data, max_items: int = 10) -> str:
        """JSON 구조 요약: 키 목록 + 샘플 항목을 Markdown으로 반환합니다."""
        lines = []
        if isinstance(data, list):
            lines.append(f"**JSON 배열** — 총 {len(data)}개 항목")
            sample = data[:max_items]
            lines.append(f"\n샘플 ({min(max_items, len(data))}개):\n")
            lines.append(json.dumps(sample, ensure_ascii=False, indent=2))
        elif isinstance(data, dict):
            lines.append(f"**JSON 객체** — 최상위 키 {len(data)}개")
            lines.append("\n키 목록:")
            for k, v in list(data.items())[:max_items]:
                v_type = type(v).__name__
                if isinstance(v, list):
                    v_repr = f"Array({len(v)})"
                elif isinstance(v, dict):
                    v_repr = f"Object({len(v)} keys)"
                else:
                    v_repr = repr(v)[:80]
                lines.append(f"- `{k}` ({v_type}): {v_repr}")
            if len(data) > max_items:
                lines.append(f"- ... 외 {len(data) - max_items}개 키")
        else:
            lines.append(json.dumps(data, ensure_ascii=False, indent=2))
        return "\n".join(lines)

    def _analyze_json_from_bytes(self, file_bytes: bytes) -> Dict[str, Any]:
        """JSON 바이트 데이터 분석 (인코딩 자동 감지 + 대용량 스키마 요약)"""
        try:
            # 인코딩 순차 시도
            encodings = ["utf-8", "utf-8-sig", "cp949", "latin-1"]
            text = None
            for enc in encodings:
                try:
                    text = file_bytes.decode(enc)
                    break
                except UnicodeDecodeError:
                    continue

            if text is None:
                return {"success": False, "error": "JSON 파일 인코딩을 감지할 수 없습니다."}

            data = json.loads(text)

            # 5KB 초과 시 스키마 요약 모드로 전환
            SUMMARY_THRESHOLD = 5_000
            raw_text = json.dumps(data, ensure_ascii=False, indent=2)
            if len(raw_text) > SUMMARY_THRESHOLD:
                json_text = self._json_schema_summary(data)
                summarized = True
            else:
                json_text = raw_text
                summarized = False

            return {
                "success": True,
                "file_type": "json",
                "text": json_text,
                "summarized": summarized,
                "word_count": len(json_text.split()),
                "char_count": len(json_text),
                "preview": json_text[:500] + "..." if len(json_text) > 500 else json_text
            }

        except Exception as e:
            return {
                "success": False,
                "error": f"JSON 분석 오류: {str(e)}"
            }
    
    def _docx_get_list_formats(self, doc) -> dict:
        """numbering.xml에서 numId별 각 레벨의 포맷(bullet/decimal 등)을 파싱합니다."""
        from docx.oxml.ns import qn
        fmt_map = {}  # {numId: {ilvl: 'bullet'|'decimal'|...}}
        try:
            numbering_part = doc.part.numbering_part
            if numbering_part is None:
                return fmt_map
            numbering_el = numbering_part.element

            # abstractNumId → {ilvl: numFmt} 매핑
            abstract_map = {}
            for abstract_num in numbering_el.findall(qn('w:abstractNum')):
                abstract_id = abstract_num.get(qn('w:abstractNumId'))
                lvl_map = {}
                for lvl in abstract_num.findall(qn('w:lvl')):
                    ilvl = lvl.get(qn('w:ilvl'), "0")
                    num_fmt_el = lvl.find(qn('w:numFmt'))
                    num_fmt = num_fmt_el.get(qn('w:val'), "bullet") if num_fmt_el is not None else "bullet"
                    lvl_map[ilvl] = num_fmt
                abstract_map[abstract_id] = lvl_map

            # num → abstractNum 매핑
            for num in numbering_el.findall(qn('w:num')):
                num_id = num.get(qn('w:numId'))
                abstract_ref = num.find(qn('w:abstractNumId'))
                if abstract_ref is not None:
                    abstract_id = abstract_ref.get(qn('w:val'))
                    fmt_map[num_id] = abstract_map.get(abstract_id, {})
        except Exception:
            pass
        return fmt_map

    def _analyze_docx_from_bytes(self, file_bytes: bytes) -> Dict[str, Any]:
        """DOCX 바이트 데이터 분석 (단락 + 표 + 리스트 순서 유지, Markdown 변환)"""
        try:
            from docx import Document as DocxDocument
            from docx.oxml.ns import qn
            from collections import defaultdict

            doc = DocxDocument(io.BytesIO(file_bytes))
            list_formats = self._docx_get_list_formats(doc)

            lines = []
            heading_count = 0
            table_count = 0
            # 번호 리스트 카운터: {(numId, ilvl): count}
            num_counters = defaultdict(int)
            prev_num_id = None  # 리스트 연속성 추적

            body = doc.element.body
            for child in body:
                tag = child.tag.split('}')[-1] if '}' in child.tag else child.tag

                if tag == 'p':
                    style_name = ""
                    pPr = child.find(qn('w:pPr'))
                    num_pr = None
                    if pPr is not None:
                        pStyle = pPr.find(qn('w:pStyle'))
                        if pStyle is not None:
                            style_name = pStyle.get(qn('w:val'), "")
                        num_pr = pPr.find(qn('w:numPr'))

                    # 텍스트 추출
                    para_text = "".join(
                        node.text or ""
                        for node in child.iter(qn('w:t'))
                    ).strip()

                    if not para_text:
                        # 리스트가 끊겼을 때 카운터 리셋
                        if prev_num_id is not None:
                            prev_num_id = None
                        lines.append("")
                        continue

                    style_lower = style_name.lower()

                    # 리스트 항목 감지 (numPr 존재)
                    if num_pr is not None:
                        ilvl_el = num_pr.find(qn('w:ilvl'))
                        num_id_el = num_pr.find(qn('w:numId'))
                        ilvl = ilvl_el.get(qn('w:val'), "0") if ilvl_el is not None else "0"
                        num_id = num_id_el.get(qn('w:val'), "0") if num_id_el is not None else "0"

                        indent = "  " * int(ilvl)
                        lvl_fmt = list_formats.get(num_id, {}).get(ilvl, "bullet")

                        # 다른 numId로 바뀌면 카운터 리셋
                        if num_id != prev_num_id:
                            num_counters.clear()
                        prev_num_id = num_id

                        if lvl_fmt == "decimal":
                            num_counters[(num_id, ilvl)] += 1
                            lines.append(f"{indent}{num_counters[(num_id, ilvl)]}. {para_text}")
                        else:
                            lines.append(f"{indent}- {para_text}")

                    # Heading 스타일 → Markdown 헤더
                    elif "heading1" in style_lower or style_name == "1":
                        lines.append(f"# {para_text}")
                        heading_count += 1
                        prev_num_id = None
                    elif "heading2" in style_lower or style_name == "2":
                        lines.append(f"## {para_text}")
                        heading_count += 1
                        prev_num_id = None
                    elif "heading3" in style_lower or style_name == "3":
                        lines.append(f"### {para_text}")
                        heading_count += 1
                        prev_num_id = None
                    elif "heading" in style_lower:
                        lines.append(f"#### {para_text}")
                        heading_count += 1
                        prev_num_id = None
                    else:
                        lines.append(para_text)
                        prev_num_id = None

                elif tag == 'tbl':
                    # 표 처리 → Markdown 파이프 테이블
                    table_count += 1
                    prev_num_id = None
                    rows = child.findall(qn('w:tr'))
                    if not rows:
                        continue

                    md_rows = []
                    for row_idx, row in enumerate(rows):
                        cells = row.findall(qn('w:tc'))
                        cell_texts = []
                        for cell in cells:
                            cell_text = "".join(
                                node.text or ""
                                for node in cell.iter(qn('w:t'))
                            ).strip().replace("|", "\\|")
                            cell_texts.append(cell_text)
                        md_rows.append("| " + " | ".join(cell_texts) + " |")
                        if row_idx == 0:
                            md_rows.append("|" + "|".join(["---"] * len(cell_texts)) + "|")

                    lines.append("")
                    lines.extend(md_rows)
                    lines.append("")

            text = "\n".join(lines).strip()
            return {
                "success": True,
                "file_type": "docx",
                "text": text,
                "word_count": len(text.split()),
                "char_count": len(text),
                "heading_count": heading_count,
                "table_count": table_count,
                "preview": text[:500] + "..." if len(text) > 500 else text
            }

        except Exception as e:
            return {
                "success": False,
                "error": f"DOCX 분석 오류: {str(e)}"
            }

    def _analyze_docx(self, file_path: str) -> Dict[str, Any]:
        """DOCX 파일 경로로 분석"""
        try:
            with open(file_path, "rb") as f:
                return self._analyze_docx_from_bytes(f.read())
        except Exception as e:
            return {
                "success": False,
                "error": f"DOCX 파일 분석 오류: {str(e)}"
            }

    def _analyze_pdf(self, file_path: str) -> Dict[str, Any]:
        """PDF 파일 분석 (품질 파이프라인 적용 버전에 위임)"""
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"PDF 파일을 찾을 수 없습니다: {file_path}")
            with open(file_path, "rb") as f:
                return self._analyze_pdf_from_bytes(f.read())
        except Exception as e:
            return {
                "success": False,
                "error": f"PDF 파일 분석 오류: {str(e)}"
            }
    
    def _analyze_excel(self, file_path: str) -> Dict[str, Any]:
        """Excel 파일 분석"""
        try:
            with open(file_path, "rb") as f:
                return self._analyze_excel_from_bytes(f.read())
        except Exception as e:
            return {
                "success": False,
                "error": f"Excel 파일 분석 오류: {str(e)}"
            }
    
    def _analyze_csv(self, file_path: str) -> Dict[str, Any]:
        """CSV 파일 분석"""
        try:
            with open(file_path, "rb") as f:
                return self._analyze_csv_from_bytes(f.read())
        except Exception as e:
            return {
                "success": False,
                "error": f"CSV 파일 분석 오류: {str(e)}"
            }
    
    def _analyze_text(self, file_path: str) -> Dict[str, Any]:
        """텍스트 파일 분석"""
        try:
            # 다양한 인코딩으로 시도
            encodings = ["utf-8", "cp949", "euc-kr", "latin-1"]
            text = ""
            
            for encoding in encodings:
                try:
                    with open(file_path, "r", encoding=encoding) as f:
                        text = f.read()
                    break
                except UnicodeDecodeError:
                    continue
            
            if not text:
                raise ValueError("텍스트 파일을 읽을 수 없습니다.")
            
            return {
                "success": True,
                "file_type": "text",
                "text": text.strip(),
                "word_count": len(text.split()),
                "char_count": len(text),
                "line_count": len(text.split("\n")),
                "preview": text[:500] + "..." if len(text) > 500 else text
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": f"텍스트 파일 분석 오류: {str(e)}"
            }
    
    def _analyze_json(self, file_path: str) -> Dict[str, Any]:
        """JSON 파일 분석"""
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            
            # JSON을 텍스트로 변환
            json_text = json.dumps(data, ensure_ascii=False, indent=2)
            
            return {
                "success": True,
                "file_type": "json",
                "text": json_text,
                "data": data,
                "word_count": len(json_text.split()),
                "char_count": len(json_text),
                "preview": json_text[:500] + "..." if len(json_text) > 500 else json_text
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": f"JSON 파일 분석 오류: {str(e)}"
            }
    
    def get_supported_formats(self) -> list:
        """지원하는 파일 형식 목록 반환"""
        return list(self.supported_formats.keys())
    
    def get_file_info(self, file_path: str) -> Dict[str, Any]:
        """파일 정보 반환"""
        try:
            stat = os.stat(file_path)
            return {
                "file_name": os.path.basename(file_path),
                "file_size": stat.st_size,
                "file_size_mb": round(stat.st_size / (1024 * 1024), 2),
                "extension": os.path.splitext(file_path)[1].lower().lstrip("."),
                "is_supported": os.path.splitext(file_path)[1].lower().lstrip(".") in self.supported_formats
            }
        except Exception as e:
            return {
                "error": f"파일 정보 조회 오류: {str(e)}"
            }
